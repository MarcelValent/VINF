# Importing necessary libraries
import lucene
from java.nio.file import Paths
from org.apache.lucene.analysis.standard import StandardAnalyzer
from org.apache.lucene.document import Document, Field, TextField
from org.apache.lucene.index import IndexWriter, IndexWriterConfig
from org.apache.lucene.store import NIOFSDirectory
import openpyxl
import pandas as pd

# Function to create an index for football matches
def create_index_matches(index_dir, excel_path):
    # Initializing Lucene components
    analyzer = StandardAnalyzer()
    config = IndexWriterConfig(analyzer)
    config.setOpenMode(IndexWriterConfig.OpenMode.CREATE)

    # Creating an index writer
    fs_directory = NIOFSDirectory(Paths.get(index_dir))
    writer = IndexWriter(fs_directory, config)

    # Loading Excel workbook
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    # Iterating over rows in the Excel sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        doc = Document()

        # Adding fields to the document
        doc.add(TextField("HomeTeam", str(row[0]), Field.Store.YES))
        doc.add(TextField("AwayTeam", str(row[1]), Field.Store.YES))
        doc.add(TextField("Date", str(row[2]), Field.Store.YES))
        doc.add(TextField("Time", str(row[3]), Field.Store.YES))

        # Handling optional and missing values for the "Stadium" field
        stadium_value = str(row[4]) if row[4] is not None and not pd.isna(row[4]) else ""
        doc.add(TextField("Stadium", stadium_value, Field.Store.YES))

        # Adding additional match-related fields
        doc.add(TextField("GoalsHome", str(row[5]), Field.Store.YES))
        doc.add(TextField("GoalsAway", str(row[6]), Field.Store.YES))
        doc.add(TextField("Goals", str(row[7]), Field.Store.YES))
        doc.add(TextField("Cards", str(row[8]), Field.Store.YES))
        doc.add(TextField("RedCards", str(row[9]), Field.Store.YES))

        # Adding the document to the index
        writer.addDocument(doc)
        # print(doc)

    # Committing changes and closing the writer
    writer.commit()
    writer.close()

# Function to create an index for football players
def create_index_players(index_dir, excel_path):
    # Initializing Lucene components
    analyzer = StandardAnalyzer()
    config = IndexWriterConfig(analyzer)
    config.setOpenMode(IndexWriterConfig.OpenMode.CREATE)

    # Creating an index writer
    fs_directory = NIOFSDirectory(Paths.get(index_dir))
    writer = IndexWriter(fs_directory, config)

    # Loading Excel workbook
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    # Iterating over rows in the Excel sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        doc = Document()

        # Adding fields to the document
        doc.add(TextField("name", str(row[0]), Field.Store.YES))
        doc.add(TextField("team", str(row[1]), Field.Store.YES))
        doc.add(TextField("position", str(row[2]), Field.Store.YES))
        doc.add(TextField("birth_date", str(row[3]), Field.Store.YES))
        doc.add(TextField("age", str(row[4]), Field.Store.YES))

        # Adding additional player-related fields
        doc.add(TextField("seasons", str(row[5]), Field.Store.YES))
        doc.add(TextField("matches", str(row[6]), Field.Store.YES))
        doc.add(TextField("goals", str(row[7]), Field.Store.YES))
        doc.add(TextField("yellow_cards", str(row[8]), Field.Store.YES))
        doc.add(TextField("red_cards", str(row[9]), Field.Store.YES))

        # Adding the document to the index
        writer.addDocument(doc)
        # print(doc)

    # Committing changes and closing the writer
    writer.commit()
    writer.close()

# Entry point of the script
if __name__ == "__main__":
    # Initializing Lucene VM
    lucene.initVM()

    # Setting file paths and directories for match and player indices
    index_directory_matches = "./index"
    excel_file_path_matches = "./matches.xlsx"
    index_directory_players = "./index_players"
    excel_file_path_players = "./players.xlsx"

    # Creating player and match indices
    create_index_players(index_directory_players, excel_file_path_players)
    create_index_matches(index_directory_matches, excel_file_path_matches)

    # Printing a message to indicate the completion of indexing
    print("Indexing is done.")

